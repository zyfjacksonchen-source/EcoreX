from __future__ import annotations

import argparse
import json
import sys
import tempfile
from pathlib import Path
from typing import Any, Dict

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from common.office_pdf_runtime import (
    OfficePdfRuntimeError,
    build_spreadsheet_quality_evidence,
    render_spreadsheet_preview,
)


def _write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _render_metadata(sheet_count: int) -> list[Dict[str, Any]]:
    return [
        {"page": sheet, "artifactRef": f"hmac:spreadsheet-render-{sheet:03d}", "width": 1280, "height": 720}
        for sheet in range(1, sheet_count + 1)
    ]


def _create_clean_workbook(path: Path) -> None:
    import openpyxl
    from openpyxl.chart import BarChart, Reference

    workbook = openpyxl.Workbook()
    inputs = workbook.active
    inputs.title = "Private Inputs"
    inputs.append(["Product", "Units", "Price", "Revenue"])
    inputs.append(["Private Alpha", 10, 2.5, "=B2*C2"])
    inputs.append(["Private Beta", 12, 3.0, "=B3*C3"])
    inputs.append(["Private Gamma", 8, 4.5, "=B4*C4"])

    dashboard = workbook.create_sheet("Private Dashboard")
    dashboard["A1"] = "Private revenue dashboard"
    dashboard["B2"] = "Total"
    dashboard["C2"] = "=SUM('Private Inputs'!D2:D4)"
    chart = BarChart()
    chart.title = "Revenue by product"
    data = Reference(inputs, min_col=4, min_row=1, max_row=4)
    categories = Reference(inputs, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    dashboard.add_chart(chart, "E2")
    workbook.save(path)


def _create_bad_workbook(path: Path) -> None:
    import openpyxl
    from openpyxl.chart import BarChart

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Private Bad Sheet"
    sheet["A1"] = "100"
    sheet["B1"] = "2026-01-01"
    sheet["C1"] = "=#REF!"
    sheet["D1"] = "Private bad workbook"
    sheet.add_chart(BarChart(), "F2")
    workbook.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Smoke R24-06 spreadsheet quality evidence.")
    parser.add_argument("--output", default="docs/v0.2.4/artifacts/spreadsheet-quality-smoke.json")
    args = parser.parse_args()

    with tempfile.TemporaryDirectory(prefix="ecorex-xlsx-quality-") as tmp:
        root = Path(tmp)
        clean_path = root / "clean-private-workbook.xlsx"
        bad_path = root / "bad-private-workbook.xlsx"
        _create_clean_workbook(clean_path)
        _create_bad_workbook(bad_path)

        render_backend = "provided-redacted-metadata"
        actual_render = None
        actual_render_available = False
        try:
            actual_render = render_spreadsheet_preview(clean_path, root / "rendered", max_sheets=2)
            render_backend = str(actual_render.get("renderBackend") or "libreoffice-pdf")
            render_items = actual_render.get("artifacts") or []
            actual_render_available = True
        except OfficePdfRuntimeError as exc:
            render_items = _render_metadata(2)
            actual_render = {"status": "unavailable", "errorType": type(exc).__name__}

        clean = build_spreadsheet_quality_evidence(
            clean_path,
            renders=render_items,
            visual_inspection_passed=True,
        )
        bad = build_spreadsheet_quality_evidence(
            bad_path,
            renders=[],
            visual_inspection_passed=True,
        )

        required_bad_failures = {"typed-values", "formula-audit", "chart-render", "render-preview"}
        expected_clean_no_render_failures = {"chart-render", "render-preview"}
        bad_failed_checks = {item["id"] for item in bad.get("checks", []) if item.get("status") == "fail"}
        clean_failed_checks = {item["id"] for item in clean.get("checks", []) if item.get("status") == "fail"}
        synthetic_render_rejected = (
            not actual_render_available
            and "render-preview" in clean_failed_checks
            and not clean.get("renderedArtifacts")
        )
        clean_render_contract_ok = (
            clean.get("status") == "pass" and not clean_failed_checks
            if actual_render_available
            else clean.get("status") == "fail"
            and synthetic_render_rejected
            and clean_failed_checks == expected_clean_no_render_failures
        )

        serialized = json.dumps({"clean": clean, "bad": bad}, ensure_ascii=False)
        leaks = [
            item
            for item in (
                "Private Alpha",
                "Private revenue dashboard",
                "Private bad workbook",
                "Private Inputs",
                "Private Dashboard",
                "Private Bad Sheet",
                str(root),
                clean_path.name,
                bad_path.name,
                "#REF!",
            )
            if item and item in serialized
        ]
        payload = {
            "status": (
                "PASS"
                if (
                    clean_render_contract_ok
                    and bad.get("status") == "fail"
                    and required_bad_failures <= bad_failed_checks
                    and not leaks
                )
                else "FAIL"
            ),
            "renderBackend": render_backend,
            "actualRender": actual_render,
            "actualRenderAvailable": actual_render_available,
            "syntheticRenderRejected": synthetic_render_rejected,
            "cleanStatus": clean.get("status"),
            "cleanFailedChecks": sorted(clean_failed_checks),
            "expectedCleanNoRenderFailures": sorted(expected_clean_no_render_failures),
            "badStatus": bad.get("status"),
            "badFailedChecks": sorted(bad_failed_checks),
            "requiredBadFailures": sorted(required_bad_failures),
            "cleanSummary": clean.get("spreadsheetAnalysis", {}).get("summary", {}),
            "badSummary": bad.get("spreadsheetAnalysis", {}).get("summary", {}),
            "leakCount": len(leaks),
            "redacted": True,
        }
        _write_json(Path(args.output), payload)
        print(json.dumps(payload, ensure_ascii=False, indent=2))
        return 0 if payload["status"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
